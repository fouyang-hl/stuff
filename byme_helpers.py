from NEW_helpers import *
from gsheet_gdrive_api import *

import pandas as pd
import numpy as np
import openpyxl as opx
from datetime import datetime
import calendar

import xlrd


pd.options.mode.chained_assignment = None  # default='warn'

########################################################
# Print Drugs CPUV campaigns that hit the goal
########################################################

def drugs_cpuv_hit_goal(path_drugs_cpuv):

    def add_sheet_summary(partner, path_partner_xlsx, list_summary):
        yesterday_date = (datetime.now() - timedelta(days=1)).date()

        partner_wb = opx.load_workbook(path_partner_xlsx, data_only=True)
        list_partner_sheets = partner_wb.get_sheet_names()
        for sheet in list_partner_sheets:
            # Skip non-data sheets
            if sheet == 'SupermetricsQueries':
                continue

            ws = partner_wb.get_sheet_by_name(sheet)
            if ws.sheet_state == 'hidden':
                hidden = 'Y'
            else:
                hidden = None

            per_sheet = []
            r = 1

            # Skip if the first date is not this month's date
            month_first_date = yesterday_date.replace(day=1)
            if (ws.cell(row=6, column=2).value.date() != month_first_date):
                list_summary.append([partner, sheet, 'Dates not in this month.', None, None, None, hidden])
                continue

            # Get to Dates rows
            while not isinstance((ws.cell(row=r, column=2).value), datetime):
                r += 1

            # Add Yesterday UVs
            while (ws.cell(row=r, column=2).value.date() != yesterday_date):
                r += 1
            per_sheet += [partner, sheet, ws.cell(row=r, column=3).value]

            # Add MTD Total
            while ((ws.cell(row=r, column=2).value != 'Total') & (ws.cell(row=r, column=2).value != 'Total MTD')):
                r += 1
            per_sheet += [ws.cell(row=r, column=3).value]

            # Add Goal
            while (ws.cell(row=r, column=2).value != 'CPUV Goal'):
                r += 1
            per_sheet += [ws.cell(row=r, column=3).value]

            # Add Need Per Day
            while (ws.cell(row=r, column=2).value != 'Daily UVs needed'):
                r += 1
            per_sheet += [ws.cell(row=r, column=3).value]

            # Sheet is hidden or not
            per_sheet += [hidden]

            list_summary.append(per_sheet)

    uvs_summary = []
    add_sheet_summary('Drugs.com', path_drugs_cpuv, uvs_summary)
    uvs_summary_df = pd.DataFrame(uvs_summary,
                                  columns=['Site', 'Report Tab Name', 'Yesterday', 'MTD', 'Goal', 'Per Day Needed',
                                           'Hidden Sheet?'])

    camps_hit_goal = uvs_summary_df[(uvs_summary_df['Goal'] > 0) &
                                    (uvs_summary_df['MTD'] >= uvs_summary_df['Goal'])]['Report Tab Name'].tolist()

    for camp in camps_hit_goal:
        print(camp)

    return None

########################################################
# Update PAS Google Sheet with Salesforce data for a given month
# das_month in the form of month/year
# use_sheet is the name of the most recent tab in PAS
########################################################

def update_pas(mo_year, use_sheet):

    ###########################################
    # Prep
    ###########################################

    mo, year = mo_year

    spreadsheetId = PAS_WHICH_GSHEET[year]
    das_month = str(mo) + '/' + str(year)

    ###########################################
    # Data to upload
    ###########################################

    # Import DAS data to use
    das = make_das(use_scheduled_units=False, export=True)
    das = das[(das[das_month] > 0) & (das['Price Calculation Type'] == 'CPM') &
              (das['Stage'] != 'Cancelled Without Activity') &
              (das['Campaign Manager'] != 'SEM')]
    das = das[['Campaign Name', 'Product', 'Campaign Manager', 'Account Name', 'Start Date', 'End Date', 'Contracted Sites', 'Line Item Number',
               'Line Description', 'Contracted Sizes', 'Sales Price', das_month]]

    # Import PAS data to use
    pas = get_pas(year, use_sheet)
    pas_header = pas.columns.tolist()
    pas_sites = pas_header[pas_header.index('Drugs'): pas_header.index('HL')+1]
    pas_use_sheet_mo_yr = pas_header[pas_header.index('Drugs Rate') + 1]

    pas = pas[['Campaign Name', 'Line Description', 'MTD Disc', 'Overall MTD Disc', 'Drugs Rate', pas_use_sheet_mo_yr, 'Drugs Rev'] + pas_sites]
    pas['Line Description'] = [ld.strip() for ld in pas['Line Description']]

    # Make updates
    if das_month == pas_use_sheet_mo_yr:
        pas_use_sheet_mo_yr = 'Old ' + pas_use_sheet_mo_yr
        pas = pas.rename(columns={das_month: pas_use_sheet_mo_yr})

    pas = pd.merge(das, pas, how='left', on=['Campaign Name', 'Line Description'])
    pas['Change'] = pas[das_month] - pas[pas_use_sheet_mo_yr]
    pas.loc[pd.isnull(pas[pas_use_sheet_mo_yr]), 'Change'] = 'New'

    pas.loc[(pas['Contracted Sites'] == 'Healthline') & (pd.isnull(pas['Drugs'])), 'Drugs'] = 'HL only'
    pas.loc[pas['Sales Price'] == 0, 'Drugs'] = 'HL only'

    pas.loc[pas['Product'].str.contains('CPM Regular-'), 'Product'] = pas[pas['Product'].str.contains('CPM Regular-')].apply(lambda row: row['Product'].replace('CPM Regular-', ''), axis=1)

    # Sort
    pas = pas[['Campaign Name', 'Product', 'Campaign Manager', 'Account Name', 'Start Date', 'End Date', 'Contracted Sites', 'Line Item Number',
               'Line Description', 'Contracted Sizes', 'MTD Disc', 'Overall MTD Disc', 'Sales Price', 'Drugs Rate',
               das_month, pas_use_sheet_mo_yr, 'Change'] + pas_sites + ['Drugs Rev']]

    paid = pas[pas['Sales Price'] > 0].sort_values(['Campaign Name', 'Line Item Number', 'Line Description'])
    av = pas[pas['Sales Price'] == 0].sort_values(['Campaign Name', 'Line Item Number', 'Line Description'])
    pas = pd.concat([paid, av]).reset_index(drop=True)

    # Formulas
    header = pas.columns.tolist()

    sales_price_col_letter = opx.utils.get_column_letter(header.index('Sales Price') + 1)
    das_month_col_letter = opx.utils.get_column_letter(header.index(das_month) + 1)
    drugs_col_letter = opx.utils.get_column_letter(header.index('Drugs') + 1)
    before_hl_col_letter = opx.utils.get_column_letter(header.index('HL'))
    drugs_rate_col_letter = opx.utils.get_column_letter(header.index('Drugs Rate') + 1)

    pas['Drugs Rate'] = [('=' + sales_price_col_letter + str(i+3) + '*0.6') for i in range(len(pas))]
    pas['HL'] = [('=' + das_month_col_letter + str(i+3) + '-SUM(' + drugs_col_letter + str(i+3) + ':' + before_hl_col_letter + str(i+3) + ')') for i in range(len(pas))]
    pas['Drugs Rev'] = [('=IFERROR(' + drugs_col_letter + str(i+3) + '/1000*' + drugs_rate_col_letter + str(i+3) + ', 0)') for i in range(len(pas))]

    # Top row
    max_row = len(pas) + 2
    top_row = []
    for i in range(len(header)):
        if i < header.index(das_month):
            top_row.append(das_month)
        else:
            col_letter = opx.utils.get_column_letter(i + 1)
            top_row.append('=SUBTOTAL(9,' + col_letter + '3:' + col_letter + str(max_row) + ')')

    # Google Sheet upload isn't doable with date type
    pas['Start Date'] = [str(d) for d in pas['Start Date']]
    pas['End Date'] = [str(d) for d in pas['End Date']]

    # Google Sheet upload isn't doable with None
    pas = pas.fillna('')

    # List of lists
    pas_values = [top_row] + [pas.columns.tolist()] + pas.values.tolist()

    # Make values json serializable
    for i in range(len(pas_values)):
        for j in range(len(pas_values[i])):
            value = pas_values[i][j]
            if isinstance(value, np.int64):
                pas_values[i][j] = int(value)
                continue
            elif isinstance(value, np.float64):
                pas_values[i][j] = float(value)
                continue

    ###########################################
    # Upload data
    ###########################################

    service = get_gsheet_service()

    # If for same month, add 'Old' to current sheet name
    abbr_month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    create_sheet = abbr_month_names[int(das_month.split('/')[0]) - 1]

    ## If 'Old' sheet exists, delete it
    gsheet_delete_sheet('Old ' + create_sheet, spreadsheetId)

    if use_sheet == create_sheet:
        ss_metadata = service.spreadsheets().get(spreadsheetId=spreadsheetId).execute()
        for sheet_metadata in ss_metadata['sheets']:
            if sheet_metadata['properties']['title'] == use_sheet:
                use_sheet_id = sheet_metadata['properties']['sheetId']
                request = [{'updateSheetProperties': {'properties': {'sheetId': use_sheet_id,
                                                                     'title': 'Old ' + use_sheet},
                                                      'fields': 'title'}}]
                result = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId,
                                                            body={'requests': request}).execute()

    # Create new sheet and upload values
    # Reference: https://developers.google.com/sheets/api/samples/sheet#add_a_sheet
    request_body = {'requests': [{'addSheet': {'properties': {'title': create_sheet,
                                                              'gridProperties': {'rowCount': len(pas_values),
                                                                                 'columnCount': len(pas_values[0])}}}}]}
    result = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=request_body).execute()
    create_sheet_id = result['replies'][0]['addSheet']['properties']['sheetId']

    create_sheet_range = create_sheet + '!A1:' + opx.utils.get_column_letter(len(pas_values[0])) + str(len(pas_values))
    result = service.spreadsheets().values().update(spreadsheetId=spreadsheetId, range=create_sheet_range,
                                                    valueInputOption='USER_ENTERED', body={'values': pas_values}).execute()
    gsheet_move_sheet(create_sheet, spreadsheetId)  # Move to far left

    ###########################################
    # Upload formatting
    ###########################################

    # Value
    value_formatting = []

    def make_vf_dict(i_row_start, n_row, i_col_start, n_col, type, pattern):
        return {'repeatCell': {'range': {'sheetId': create_sheet_id,
                               'startRowIndex': i_row_start,
                               'endRowIndex': i_row_start + n_row,
                               'startColumnIndex': i_col_start,
                               'endColumnIndex': i_col_start + n_col},
                               'cell': {'userEnteredFormat': {'numberFormat': {'type': type,
                                                                               'pattern': pattern}}},
                               'fields': 'userEnteredFormat.numberFormat'}}

    value_formatting.append(make_vf_dict(2, len(pas_values) - 2, header.index('Start Date'), 2, 'DATE', 'm/d/yyyy'))
    value_formatting.append(make_vf_dict(2, len(pas_values) - 2, header.index('MTD Disc'), 2, 'NUMBER', '#0.0%'))
    value_formatting.append(make_vf_dict(2, len(pas_values) - 2, header.index('Sales Price'), 2, 'NUMBER', '$####0.00'))
    value_formatting.append(make_vf_dict(0, len(pas_values), header.index('Drugs Rev'), 1, 'NUMBER', '$#,###,##0'))
    value_formatting.append(make_vf_dict(2, len(pas_values) - 2, header.index(das_month), len(header) - header.index(das_month) - 1, 'NUMBER', '###,###,###'))
    value_formatting.append(make_vf_dict(0, 1, header.index(das_month), len(header) - header.index(das_month) - 1, 'NUMBER', '###,###,###'))

    # Color
    default = {'red': 1, 'green': 1, 'blue': 1, 'alpha': 1}
    blue = {'red': .812, 'green': .886, 'blue': .953, 'alpha': 1}
    green = {'red': .851, 'green': .918, 'blue': .827, 'alpha': 1}
    pink = {'red': 1, 'green': .8555, 'blue': .957, 'alpha': 1}
    yellow = {'red': 1, 'green': .949, 'blue': .8, 'alpha': 1}

    def repeat_color(color, n):
        return [{'userEnteredFormat': {'backgroundColor': color}}] * n

    top_row_colors = []
    top_row_colors += repeat_color(pink, header.index(das_month))
    top_row_colors += repeat_color(yellow, 3)
    top_row_colors += repeat_color(green, 1)
    top_row_colors += repeat_color(blue, len(pas_sites) - 1)
    top_row_colors += repeat_color(green, 1)

    header_colors = []
    header_colors += repeat_color(blue, header.index('MTD Disc'))
    header_colors += repeat_color(green, 1)
    header_colors += repeat_color(blue, 2)
    header_colors += repeat_color(green, 1)
    header_colors += repeat_color(yellow, 3)
    header_colors += repeat_color(green, 1)
    header_colors += repeat_color(pink, len(pas_sites) - 1)
    header_colors += repeat_color(green, 1)

    data_colors = []
    data_colors += repeat_color(default, header.index('MTD Disc'))
    data_colors += repeat_color(green, 1)
    data_colors += repeat_color(default, 2)
    data_colors += repeat_color(green, 1)
    data_colors += repeat_color(yellow, 3)
    data_colors += repeat_color(green, 1)
    data_colors += repeat_color(default, len(pas_sites) - 1)
    data_colors += repeat_color(green, 1)

    colors = []
    colors += [{'values': top_row_colors}]
    colors += [{'values': header_colors}]
    colors += [{'values': data_colors}] * (len(pas_values) - 2)

    color_formatting = [{'updateCells': {'rows': colors,
                                         'fields': 'userEnteredFormat.backgroundColor',
                                         'range': {'sheetId': create_sheet_id,
                                                   'startRowIndex': 0,
                                                   'endRowIndex': len(pas_values) + 1,
                                                   'startColumnIndex': 0,
                                                   'endColumnIndex': len(pas_values[0]) + 1}}}]

    # Freeze
    # Reference: https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets#gridproperties

    freeze = [{'updateSheetProperties': {'properties': {'sheetId': create_sheet_id,
                                                        'gridProperties': {'frozenRowCount': 2,
                                                                           'frozenColumnCount': 3}},
                                         'fields': 'gridProperties(frozenRowCount, frozenColumnCount)'}}]

    # Wrap
    clip = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'wrapStrategy': 'CLIP'}}]}] * (len(pas_values) - 2),
                             'fields': 'userEnteredFormat.wrapStrategy',
                             'range': {'sheetId': create_sheet_id,
                                       'startRowIndex': 2,
                                       'endRowIndex': len(pas_values) + 1,
                                       'startColumnIndex': header.index('Contracted Sizes'),
                                       'endColumnIndex': header.index('Contracted Sizes') + 1}}}]

    # Font
    font = [{'updateCells': {'rows': [{'values': [{'userEnteredFormat': {'textFormat': {'fontFamily': 'Fresca',
                                                                                        'fontSize': 10}}}] * len(pas_values[0])}] * len(pas_values),
                             'fields': 'userEnteredFormat.textFormat(fontFamily, fontSize)',
                             'range': {'sheetId': create_sheet_id,
                                       'startRowIndex': 0,
                                       'endRowIndex': len(pas_values) + 1,
                                       'startColumnIndex': 0,
                                       'endColumnIndex': len(pas_values[0]) + 1}}}]

    # column width
    width = [{'updateDimensionProperties': {'range': {'sheetId': create_sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': 0,
                                                      'endIndex': len(header)},
                                            'properties': {'pixelSize': 60},
                                            'fields': 'pixelSize'}},
             {'updateDimensionProperties': {'range': {'sheetId': create_sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': 0,
                                                      'endIndex': 1},
                                            'properties': {'pixelSize': 85},
                                            'fields': 'pixelSize'}},
             {'updateDimensionProperties': {'range': {'sheetId': create_sheet_id,
                                                      'dimension': 'COLUMNS',
                                                      'startIndex': header.index('Line Description'),
                                                      'endIndex': header.index('Line Description') + 1},
                                            'properties': {'pixelSize': 200},
                                            'fields': 'pixelSize'}}]

    all_formatting_requests = value_formatting + color_formatting + freeze + clip + font + width
    result = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body={'requests': all_formatting_requests}).execute()

def compare_pas(year, before_pas_sheet, now_pas_sheet):

    now_pas = get_pas(year, now_pas_sheet)
    before_pas = get_pas(year, before_pas_sheet)

    header = now_pas.columns.tolist()
    list_sites = header[header.index('Drugs'): header.index('HL')]

    now_pas = now_pas[['Campaign Name', 'Line Item Number', 'Line Description'] + list_sites]
    before_pas = before_pas[['Campaign Name', 'Line Description'] + list_sites]

    for site in list_sites:
        now_pas[site] = [0 if isinstance(unit, str) else unit for unit in now_pas[site]]
        before_pas[site] = [0 if isinstance(unit, str) else unit for unit in before_pas[site]]

    combined = pd.merge(now_pas, before_pas, how='left', on=['Campaign Name', 'Line Description'],
                        suffixes=[' (' + now_pas_sheet + ')', ' (' + before_pas_sheet + ')'])
    combined = combined.fillna(0)

    for site in list_sites:
        combined[site + ' (Diff)'] = combined[site + ' (' + now_pas_sheet + ')'] - combined[site + ' (' + before_pas_sheet + ')']

    def pick_non_zero_diff(row):
        for site in list_sites:
            if row[site + ' (Diff)'] != 0:
                return 'Changed'
        return ''

    combined['Changed?'] = combined.apply(lambda row: pick_non_zero_diff(row), axis=1)
    col = combined.columns.tolist()
    col.remove('Changed?')
    col.insert(0, 'Changed?')
    combined = combined[col]

    return combined

def compare_cpuv_goals(year, before_cpuv_goals_sheet, now_cpuv_goals_sheet):

    now_cpuv_goals = get_cpuv_goals(year, now_cpuv_goals_sheet)
    before_cpuv_goals = get_cpuv_goals(year, before_cpuv_goals_sheet)

    header = now_cpuv_goals.columns.tolist()
    list_sites = header[header.index('Drugs Goal'): header.index('Goal Check')]

    now_cpuv_goals = now_cpuv_goals[['Campaign Name', 'Line Item Number', 'Line Description'] + list_sites]
    before_cpuv_goals = before_cpuv_goals[['Campaign Name', 'Line Description'] + list_sites]

    for site in list_sites:
        now_cpuv_goals[site] = [0 if isinstance(unit, str) else unit for unit in now_cpuv_goals[site]]
        before_cpuv_goals[site] = [0 if isinstance(unit, str) else unit for unit in before_cpuv_goals[site]]

    combined = pd.merge(now_cpuv_goals, before_cpuv_goals, how='left', on=['Campaign Name', 'Line Description'],
                        suffixes=[' (' + now_cpuv_goals_sheet + ')', ' (' + before_cpuv_goals_sheet + ')'])
    combined = combined.fillna(0)

    for site in list_sites:
        combined[site + ' (Diff)'] = combined[site + ' (' + now_cpuv_goals_sheet + ')'] - combined[site + ' (' + before_cpuv_goals_sheet + ')']

    def pick_non_zero_diff(row):
        for site in list_sites:
            if row[site + ' (Diff)'] != 0:
                return 'Changed'
        return ''

    combined['Changed?'] = combined.apply(lambda row: pick_non_zero_diff(row), axis=1)
    col = combined.columns.tolist()
    col.remove('Changed?')
    col.insert(0, 'Changed?')
    combined = combined[col]

    header = combined.columns.tolist()
    header_rename_dict = {}
    for col in header:
        if ' Goal' in col:
            header_rename_dict[col] = col.replace(' Goal', '')
    combined = combined.rename(columns=header_rename_dict)

    return combined

########################################################
# Given a PAS df, return the allocated goal and
# % of the overall goal for each site per Brand/Campaign
########################################################

def get_volume_share(pas):
    header = pas.columns.tolist()
    list_sites = header[header.index('Drugs'): header.index('HL')+1]

    camp_brand = make_das(use_scheduled_units=False, export=True)[['Campaign Name', 'Brand']].drop_duplicates()
    pas = pd.merge(pas, camp_brand, how='left', on='Campaign Name')
    pas.loc[pas['Brand'] == 'Unknown/Other (Brand)', 'Brand'] = \
        pas.loc[pas['Brand'] == 'Unknown/Other (Brand)', 'Campaign Name']
    pas = pas[['Brand'] + list_sites]

    for site in list_sites:
        pas[site] = [0 if isinstance(unit, str) else unit for unit in pas[site]]
    pas = pas.fillna(0)

    vol_share = pas.groupby('Brand').sum().reset_index()
    vol_share['Total'] = vol_share.apply(lambda row: sum(row[site] for site in list_sites), axis=1)

    for site in list_sites:
        vol_share['% ' + site] = vol_share[site] / vol_share['Total']

    return vol_share

########################################################
# Given 2 PAS sheet names, return a df of Brand/Campaign
# that HL's volume share changed by 10% or more
########################################################

def vol_share_shift_over10p(year, before_pas_sheet, now_pas_sheet):

    # Get before and now volume share dfs
    before_pas = get_pas(year, before_pas_sheet)
    now_pas = get_pas(year, now_pas_sheet)

    before_vs = get_volume_share(before_pas)
    now_vs = get_volume_share(now_pas)

    # Add 'before' or 'now' to the header
    def rename_col(vs, prefix):
        rename_dict = {}
        for col in vs.columns.tolist():
            if col != 'Brand':
                rename_dict[col] = prefix + ' ' + col
        return vs.rename(columns=rename_dict)

    before_vs = rename_col(before_vs, 'before')
    now_vs = rename_col(now_vs, 'now')

    # Join before and now, and extract campaigns with over 10% change in HL volume share
    combined = pd.merge(now_vs, before_vs, how='left', on='Brand').fillna(0)
    combined['Change % HL'] = combined['now % HL'] - combined['before % HL']

    over10p = combined[[abs(change) >= 0.1 for change in combined['Change % HL']]]
    over10p = over10p[['Brand', 'Change % HL', 'now % HL', 'now % Drugs', 'before % HL']]

    return over10p

########################################################
# Update Partner IO Naming
########################################################

def update_drugs_io_naming(mo_year, pas_sheet, cpuv_goals_sheet):
    update_partner_io_naming('Drugs', mo_year)

def update_goodrx_io_naming(mo_year, pas_sheet, cpuv_goals_sheet):
    update_partner_io_naming('GoodRx', mo_year)

def update_partner_io_naming(site, mo_year):

    #################################################################
    if site == 'Drugs':
        spreadsheetId = '1Mx3F6K1jnf01ra2sjutmia7rMULLNlEacCcHjo98-j0'
    elif site == 'GoodRx':
        spreadsheetId = '1nCwy9nCzLcDqbhXHMNg8opT3zVtIWuBvpnkbFwjvbBQ'
    #################################################################

    if mo_year == 'Booked':  # Special case to update Booked tab
        site_io_naming_sheet = 'Booked'
        col = ['Campaign Name', 'Line Description', 'Price Calculation Type']
        site_has_goals = get_booked_future_months(site)[col].drop_duplicates()
    else:  # Monthly tab
        mo, year = mo_year
        site_io_naming_sheet = str(year) + str(mo).zfill(2)
        month_name = calendar.month_name[mo][: 3]

        # PAS
        pas = get_pas(year, month_name)
        site_has_goals_cpm = pas[[not isinstance(d, str) for d in pas[site]]]
        site_has_goals_cpm = site_has_goals_cpm[site_has_goals_cpm[site] > 0]
        site_has_goals_cpm = site_has_goals_cpm[['Campaign Name', 'Line Description']]
        site_has_goals_cpm['Price Calculation Type'] = 'CPM'

        # CPUV Goals Sheet
        cpuv_goals = get_cpuv_goals(year, month_name)
        site_has_goals_cpuv = cpuv_goals[[not isinstance(d, str) for d in cpuv_goals[site + ' Goal']]]
        site_has_goals_cpuv = site_has_goals_cpuv[site_has_goals_cpuv[site + ' Goal'] > 0]
        site_has_goals_cpuv = site_has_goals_cpuv[['Campaign Name', 'Line Description']]
        site_has_goals_cpuv['Price Calculation Type'] = 'CPUV'

        # Combine the two
        site_has_goals = pd.concat([site_has_goals_cpm, site_has_goals_cpuv])

    # Site IO Naming
    site_io_naming = get_partner_io_naming(site, site_io_naming_sheet)
    site_io_naming_1 = site_io_naming[['Internal Campaign Name', 'Line Description']].rename(
        columns={'Internal Campaign Name': 'Campaign Name'})
    site_io_naming_1['In site_io_naming'] = 'Y'

    site_io_naming_2 = site_io_naming[['Campaign Name', 'Internal Campaign Name']].drop_duplicates()
    site_io_naming_2 = site_io_naming_2[pd.notnull(site_io_naming_2['Internal Campaign Name'])]

    # Find what's missing in Site IO Naming
    site_has_goals = pd.merge(site_has_goals, site_io_naming_1, how='left', on=['Campaign Name', 'Line Description'])
    missing_in_site_io_naming = site_has_goals[site_has_goals['In site_io_naming'] != 'Y']
    missing_in_site_io_naming = missing_in_site_io_naming.drop('In site_io_naming', axis=1)

    # Ask for Campaign Name and Placement for Site IO. If the same campaign exists, don't ask.
    add_to_site_io_naming = []
    for i, row in missing_in_site_io_naming.iterrows():
        exist_campaign_in_site_io_naming = site_io_naming_2[site_io_naming_2['Internal Campaign Name'] == row['Campaign Name']].index.tolist()
        if len(exist_campaign_in_site_io_naming) > 0:
            campaign_drugs = site_io_naming_2.loc[exist_campaign_in_site_io_naming[0]][['Campaign Name']].values.tolist()[0]
        else:
            campaign_drugs = input('Enter [Campaign Name] "' + row['Campaign Name'] + '" for ' + site + ' IO: ')
            site_io_naming_2 = site_io_naming_2.append(pd.DataFrame([[campaign_drugs, row['Campaign Name']]],
                                            columns=['Campaign Name', 'Internal Campaign Name']))
            site_io_naming_2 = site_io_naming_2.reset_index(drop=True)
        placement_drugs = input(
            'Enter [' + row['Price Calculation Type'] + ' Placement] for ' + campaign_drugs + ' "' + row[
                'Line Description'] + '" for ' + site + ' IO: ')
        add_to_site_io_naming.append([campaign_drugs, row['Line Description'], placement_drugs, row['Campaign Name'],
                          row['Price Calculation Type']])
    add_to_site_io_naming_df = pd.DataFrame(add_to_site_io_naming,
                                columns=['Campaign Name', 'Line Description', 'Placement', 'Internal Campaign Name',
                                         'Price Calculation Type'])

    # Output
    revised_site_io_naming = pd.concat([site_io_naming, add_to_site_io_naming_df]).sort_values(
        ['Price Calculation Type', 'Campaign Name', 'Line Description'])

    site_has_goals['Has goal'] = 'Y'
    site_has_goals = site_has_goals.rename(columns={'Campaign Name': 'Internal Campaign Name'})
    site_has_goals = site_has_goals[['Internal Campaign Name', 'Line Description', 'Has goal']]

    revised_site_io_naming = pd.merge(revised_site_io_naming, site_has_goals, how='left', on=['Internal Campaign Name', 'Line Description'])
    revised_site_io_naming = revised_site_io_naming[revised_site_io_naming['Has goal'] == 'Y']
    revised_site_io_naming = revised_site_io_naming[
        ['Campaign Name', 'Line Description', 'Placement', 'Internal Campaign Name', 'Price Calculation Type']]

    # Send to Google Sheet
    service = get_gsheet_service()

    result = service.spreadsheets().values().clear(spreadsheetId=spreadsheetId, range=site_io_naming_sheet, body={}).execute()
    print("Cleared!")

    values = [revised_site_io_naming.columns.tolist()] + revised_site_io_naming.values.tolist()

    print("Starting UPLOAD")
    result = service.spreadsheets().values().update(spreadsheetId=spreadsheetId, range=site_io_naming_sheet,
                                                    valueInputOption='USER_ENTERED', body={'values': values}).execute()

    print("UPLOAD DONE")

    return None

########################################################
# YTD Monthly Drugs.com revenue
# per CPM, CPUV Microsite, CPUV CC
########################################################

def get_ytd_drugs_revenue(path_dict, just_sr=False):
    month_dict = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                  7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    combined = pd.DataFrame()
    for mo in path_dict:
        try:
            df = pd.read_excel(path_dict[mo], sheetname='Billable')
        except xlrd.biffh.XLRDError:
            df = pd.read_excel(path_dict[mo], sheetname=month_dict[mo] + ' Billable')

        df['Month'] = mo
        combined = combined.append(df)

    if just_sr:
        return combined

    def add_drugs_rev_type(row):
        if row['Site'] not in ['Drugs.com', 'Drugs.com NG']:
            return None
        if row['Site'] == 'Drugs.com NG':
            return 'NG'

        # Site is 'Drugs.com' i.e. expense counts towards guarantee
        if row['Unit'] == 'CPM':
            return 'CPM'
        if row['Unit'] == 'CPUV':
            if 'competitive conquesting' in row['Line Item Name'].lower():
                try:
                    if ('Ruconest' in row['Campaign Name']) and ('PLACEHOLDER' not in row['Line Item Name']):  # Special case
                        return 'CPUV Microsite'
                    else:
                        return 'CPUV CC'
                except TypeError:
                    return 'CPUV CC'
            elif 'brand championing' in row['Line Item Name'].lower():
                return 'CPUV CC'
            else:
                return 'CPUV Microsite'
        return 'Unknown Type'

    combined['Drugs Rev Type'] = combined.apply(lambda row: add_drugs_rev_type(row), axis=1)

    # Monthly Drugs Rev per Type
    drugs = combined[(combined['Site'] == 'Drugs.com') | (combined['Site'] == 'Drugs.com NG')]
    drugs = pd.pivot_table(drugs, values='Net Site Expense', index=['Month'], columns=['Drugs Rev Type'],
                           aggfunc=np.sum, fill_value=0)
    types = drugs.columns.tolist()
    drugs = drugs.reset_index(drop=False)

    # Add Total column
    drugs['Total'] = drugs.apply(lambda row: sum([row[col] for col in types]), axis=1)
    drugs['7.26 MM Guarantee'] = drugs.apply(lambda row: row['CPM'] + row['CPUV CC'] + row['CPUV Microsite'], axis=1)

    column_order = ['Month', '7.26 MM Guarantee'] + types + ['Total']
    drugs = drugs[column_order]

    return combined, drugs

def get_booked_future_months(site):
    # Pick up future months' goals
    revshare = get_revshare_dict()[site]

    das = make_das()
    das_columns = das.columns.tolist()
    goal_columns = []

    for col in das_columns:
        if re.search('^[0-9]+/[0-9]{4}$', col):
            # Current month and year
            current = datetime.now()
            current_year = current.year
            current_mo = current.month

            # Only pick up future months
            mo, year = [int(d) for d in col.split('/')]
            if year < current_year:
                continue
            if (year == current_year) & (mo <= current_mo):
                continue
            goal_columns.append(col)

    booked_dfs = []

    ## CPM
    for goal_col in goal_columns:
        booked_cpm = das[(das['Contracted Sites'] == site) &  # This partner site only
                         (das['Price Calculation Type'] == 'CPM') &
                         (das[goal_col] > 0)]
        booked_cpm = booked_cpm.rename(columns={goal_col: 'Booked'})
        booked_cpm['Month/Year'] = goal_col
        booked_dfs.append(booked_cpm)

    ## CPUV CC
    for goal_col in goal_columns:
        booked_cpuv = das[(das['Contracted Sites'].str.contains(site, case=False)) &  # Includes this partner site
                          (das['Price Calculation Type'] == 'CPUV') &
                          (das['Media Product'] == 'Competitive Conquesting') &
                          (das[goal_col] > 0)]
        booked_cpuv = booked_cpuv.rename(columns={goal_col: 'Booked'})
        booked_cpuv['Month/Year'] = goal_col
        booked_dfs.append(booked_cpuv)

    # Format output
    booked = pd.concat(booked_dfs)  # Combine
    booked['Base Rate'] = booked['Base Rate'] * revshare  # Net CPM/CPUV
    booked.loc[booked['Contracted Sites'] != site, 'Booked'] = 'TBD'  # If a goal is a combined goal, replace with 'TBD'

    # Clean up
    cols = ['Price Calculation Type', 'Campaign Name', 'Line Description', 'Month/Year', 'Booked', 'Base Rate']
    booked = booked[cols]

    return booked

"""
########################################################
# Drugs CPM Daily Rev per Camp, Placement, Size
########################################################

def druvs_daily_rev_per_camp_pl_size(path_nr, output):

    # Import
    das = make_das(use_scheduled_units=False, export=True)
    data = pd.read_excel(path_nr, sheetname='data1')
    info = pd.read_excel(path_nr, sheetname='data2')
    labeling = get_drugs_io_naming('201706')

    # Select rows
    data = data[(data['Price Calculation Type'] == 'CPM') &
                (data['Site'] == 'Drugs.com') &
                (pd.notnull(data['Gross Revenue']))]

    # Add Internal Campaign Name
    data = bbr2camp(data, '(DAS)BBR #', das).rename(columns={'(DAS)BBR #': 'BBR',
                                                             'Campaign Name': 'Internal Campaign Name',
                                                             'DAS Line Item Name': 'Line Description'})

    # Add net rate
    info = info[info['Site'] == 'Drugs.com'].rename(columns={'DAS Line Item Name': 'Line Description'})
    info = info[['BBR', 'Line Description', 'Site Rate']]
    data = pd.merge(data, info, how='left', on=['BBR', 'Line Description'])

    # Add Drugs IO Campaign Name and Placement
    data = pd.merge(data, labeling, how='left', on=['Price Calculation Type',
                                                    'Internal Campaign Name',
                                                    'Line Description'])

    # Select columns
    data = data[['Campaign Name', 'Placement', 'Creative size', 'Date', 'Impressions/UVs',
                 'MTD Disc', 'Site Rate']]
    data = data.rename(columns={'Impressions/UVs': 'Impressions',
                                'Creative size': 'Size',
                                'Site Rate': 'CPM'})
    data = data.sort_values(['Campaign Name', 'Placement', 'Size', 'Date'])

    # Add Revenue column
    data['Revenue'] = ['=E'+str(i+2)+'*(1-F'+str(i+2)+')/1000*G'+str(i+2) for i in range(len(data))]

    # Export to csv
    data.to_csv(output, index=False)

    return None
"""

